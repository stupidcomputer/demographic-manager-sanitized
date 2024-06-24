import json
import common
import interndata
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference
)

from openpyxl.chart.series import DataPoint

def convert_x_y_to_col_id(x, y):
    return get_column_letter(x) + str(y)

def convert_x_y_range_to_cols(x1, y1, x2, y2):
    return "{}:{}".format(convert_x_y_to_col_id(x1, y1), convert_x_y_to_col_id(x2, y2))

def mapper(string):
    # map to human-readable values
    try:
        return common.mapper_data[str(string)]
    except KeyError:
        return "WARNING: {}".format(string)

def aggregate_demographic_totals(payload, adultonly, childonly, internonly, withoutinterns):
    results = {"age": {}, "gender": {}, "ethnicity": {}, "race": {}}

    for i in payload:
        if adultonly and not i["age"] == "a":
            continue

        if childonly and not i["age"] == "c":
            continue

        if internonly and not i["age"] == "i":
            continue

        if withoutinterns and i["age"] == "i":
            continue

        # for every entry we get
        for key in results.keys():
            # for every field we have
            results[key].setdefault(i[key], 0)
            results[key][i[key]] += 1

    return results

def get_totals_for_specific_attribute_on_specific_site(site, adultonly=False, childonly=False, internonly=False, withoutinterns=False):
    manifest = open("data/manifest.json")
    manifest_data = json.loads(manifest.read())

    record_to_be_used = manifest_data["sites"]
    site_object = None
    for i in manifest_data["sites"]:
        if i["name"] == site:
            site_object = i

    if not site_object:
        raise TypeError("site argument is not valid")

    used_filename = site_object["usedrecord"]
    if not used_filename: # it's just a placeholder, so we should return no records
        return {"age": {}, "gender": {}, "ethnicity": {}, "race": {}}

    present_interns = site_object["present"]

    intern_serialized = interndata.get_staff_records()
    intern_data = interndata.return_intern_data_from_present_array(present_interns, intern_serialized)

    site_specific_data_file = open("data/" + used_filename)

    demographic_data_json = json.loads(site_specific_data_file.read())

    payload = demographic_data_json["payload"]

    return aggregate_demographic_totals(payload + intern_data, adultonly, childonly, internonly, withoutinterns)

def generate_table_for_attr(attr, totals):
    header = [mapper(attr), "Count"]
    data = []

    for key, value in totals[attr].items():
        data.append([mapper(key), value])
        data.sort()

    data.insert(0, header)

    return data

def write_dataarray_to_specific_cell(x, y, dataarray, worksheet):
    i_c = 0
    for i in dataarray:
        j_c = 0
        for j in i:
            worksheet.cell(row=i_c + 1 + x, column=j_c + 1 + y, value=dataarray[i_c][j_c])
            j_c += 1
        i_c += 1

def generate_provider_string(providers):
    used_providers = sorted(providers)
    if len(used_providers) > 1:
        before_and = ', '.join(used_providers[:-1])
        if len(used_providers) > 2:
            and_string = ', and {}'.format(used_providers[-1])
        else:
            and_string = ' and {}'.format(used_providers[-1])
        return before_and + and_string
    try:
        return used_providers[0]
    except: # there's no documented communications coordinator or whatever
        return "an Unknown Person"

def handle_writing_of_attrs(worksheet, totals, offset):
    count = 0
    maxlen = 0
    for i in totals.keys():
        tmp = generate_table_for_attr(i, totals)
        maxlen = max(maxlen, len(tmp))

        write_dataarray_to_specific_cell(offset, count * 3, tmp, worksheet)
        count += 1

    return maxlen

def adjust_column_width(worksheet, times):
    for i in range(times):
        worksheet.column_dimensions[get_column_letter((i * 3) + 1)].width = 20

def write_information_to_spreadsheet(totals, worksheet, offset):
    return handle_writing_of_attrs(worksheet, totals, offset)

def plural(arr):
    if len(arr) > 1:
        return "s"
    return ""

def handle_spreadsheet_decoration(ws, providers, persons, commleads, onsiteleads, time, date):
    ws["A1"].value = "Information for {} site attendance at {} on {}".format(ws.title, time, date)
    ws["A2"].value = "Fields not present should be assumed 0."
    ws["A3"].value = "Data submitted by {} in person after conclusion of site operations.".format(generate_provider_string(providers))
    ws["A4"].value = "Interns present: {}".format(generate_provider_string(persons))
    ws["A5"].value = "Communication Lead{}: {}".format(plural(commleads), generate_provider_string(commleads))
    ws["A6"].value = "On Site Lead{}: {}".format(plural(onsiteleads), generate_provider_string(onsiteleads))
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")
    ws.merge_cells("A3:H3")
    ws.merge_cells("A4:H4")
    ws.merge_cells("A5:H5")
    ws.merge_cells("A6:H6")

def get_all_data_points_from_all_sites(manifest):
    files_to_read = []
    record_objects = []
    for site in manifest["sites"]:
        files_to_read.append(site["usedrecord"])

    for file in files_to_read:
        handle = open("data/" + file, "r")
        handle_data = json.loads(handle.read())

        payload = handle_data["payload"]
        record_objects += payload

    return record_objects

def dict_to_2d_array(dictionary, keylambda=lambda x: x):
    return [(keylambda(k), v) for k, v in dictionary.items()]

def correct_dict_with_none_values(dictionary):
    try:
        dictionary["unk"] += dictionary[None]
        del dictionary[None]
    except KeyError:
        pass

    return dictionary

def generate_final_report_data(ws):
    fd = open("data/manifest.json", "r")
    manifest = json.loads(fd.read())

    global_data_points = get_all_data_points_from_all_sites(manifest)
    total_sites = len(manifest["sites"])

    children = []
    for person in global_data_points:
        if person["age"] == "c":
            children.append(person)

    aggregation = aggregate_demographic_totals(global_data_points, False, False, False, False)
    children_aggregation = aggregate_demographic_totals(children, False, False, False, False)

    total_children = children_aggregation["age"]["c"]
    total_girls = children_aggregation["gender"]["f"]
    total_boys = children_aggregation["gender"]["m"]
    race_breakdown = sorted(dict_to_2d_array(correct_dict_with_none_values(children_aggregation["race"]), mapper))
    ethnicity_breakdown = sorted(dict_to_2d_array(correct_dict_with_none_values(children_aggregation["ethnicity"]), mapper))

    gender_breakdown = [
        ["Gender", "Count"],
        ["Boys", total_boys],
        ["Girls", total_girls],
    ]

    race_breakdown.insert(0, ["Race", "Count"])

    final_report_data_to_append = [
        ["Final Report", ""],
        ["", ""],
        ["Gender Breakdown", ""],
        *gender_breakdown,
        ["", ""],
        ["Race Breakdown", ""],
        *race_breakdown,
        ["", ""],
        ["Ethnicity Breakdown", ""],
        *ethnicity_breakdown,
        ["", ""],
        ["Total Children Served", ""],
        [total_girls + total_boys, ""],
        ["", ""],
        ["Number of Sites", ""],
        [total_sites, ""],
        ["", ""],
        ["Average Children per Site", ""],
        [(total_girls + total_boys) / total_sites, ""],
    ]

    for row in final_report_data_to_append:
        ws.append(row)

    gender_pie = PieChart()
    data = Reference(ws, min_col=2, min_row=5, max_row=5 + len(gender_breakdown) - 2)
    labels = Reference(ws, min_col=1, min_row=5, max_row=5 + len(gender_breakdown) - 2)
    gender_pie.add_data(data)
    gender_pie.set_categories(labels)
    gender_pie.title = "Gender Breakdown"
    ws.add_chart(gender_pie, "D1")

    race_pie = PieChart()
    data = Reference(ws, min_col=2, min_row=5 + len(gender_breakdown) + 2, max_row=5 + len(gender_breakdown) + len(race_breakdown))
    labels = Reference(ws, min_col=1, min_row=5 + len(gender_breakdown) + 2, max_row=5 + len(gender_breakdown) + len(race_breakdown))
    race_pie.add_data(data)
    race_pie.set_categories(labels)
    race_pie.title = "Race Breakdown"
    ws.add_chart(race_pie, "D17")

    ethnicity_pie = PieChart()
    data = Reference(ws, min_col=2, min_row=18, max_row=19)
    labels = Reference(ws, min_col=1, min_row=18, max_row=19)
    ethnicity_pie.add_data(data)
    ethnicity_pie.set_categories(labels)
    ethnicity_pie.title = "Ethnicity Breakdown"
    ws.add_chart(ethnicity_pie, "D33")

    ws.column_dimensions["A"].width = 27

    print(children_aggregation)
    print(total_children, total_girls, total_boys)
    print(race_breakdown)
    print(final_report_data_to_append)


fd = open("data/manifest.json", "r")
json_data = json.loads(fd.read())
wb = Workbook()

for site in json_data["sites"]:
    print(site["name"])
    providers = []
    for record in site["records"]:
        providers.append(record['submitter'])

    persons = site["present"]
    commleads = site["commleads"]
    onsiteleads = site["onsiteleads"]
    time = site["time"]
    date = site["date"]

    length = 8
    ws = wb.create_sheet(site["name"])

    ws.cell(row=length, column=1, value="Combined Totals")
    ws.merge_cells(convert_x_y_range_to_cols(1, length, 11, length))
    ws[convert_x_y_to_col_id(1, length)].font = Font(italic=True)
    length += write_information_to_spreadsheet(get_totals_for_specific_attribute_on_specific_site(site["name"]), ws, length) + 2

    ws.cell(row=length, column=1, value="Combined Totals (without interns)")
    ws.merge_cells(convert_x_y_range_to_cols(1, length, 11, length))
    ws[convert_x_y_to_col_id(1, length)].font = Font(italic=True)
    length += write_information_to_spreadsheet(get_totals_for_specific_attribute_on_specific_site(site["name"], withoutinterns=True), ws, length) + 2

    ws.cell(row=length, column=1, value="Adults")
    ws.merge_cells(convert_x_y_range_to_cols(1, length, 11, length))
    ws[convert_x_y_to_col_id(1, length)].font = Font(italic=True)
    length += write_information_to_spreadsheet(get_totals_for_specific_attribute_on_specific_site(site["name"], adultonly=True), ws, length) + 3

    ws.cell(row=length, column=1, value="Children")
    ws.merge_cells(convert_x_y_range_to_cols(1, length, 11, length))
    ws[convert_x_y_to_col_id(1, length)].font = Font(italic=True)
    length += write_information_to_spreadsheet(get_totals_for_specific_attribute_on_specific_site(site["name"], childonly=True), ws, length) + 3

    ws.cell(row=length, column=1, value="Interns")
    ws.merge_cells(convert_x_y_range_to_cols(1, length, 11, length))
    ws[convert_x_y_to_col_id(1, length)].font = Font(italic=True)
    length += write_information_to_spreadsheet(get_totals_for_specific_attribute_on_specific_site(site["name"], internonly=True), ws, length) + 3

    adjust_column_width(ws, 4)
    handle_spreadsheet_decoration(ws, providers, persons, commleads, onsiteleads, time, date)

ws = wb.create_sheet("Final Report")

generate_final_report_data(ws)

del wb["Sheet"]
wb.save("test.xlsx")
